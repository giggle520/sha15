import openpyxl
from openpyxl.styles import Alignment
from datetime import datetime
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side
from openpyxl.styles import Font


def format_pace(pace_str):
    if not pace_str or ':' not in pace_str:
        return pace_str
    minutes, seconds = pace_str.split(':')
    return f"{minutes}'" + f'{seconds}"'

def pace_to_seconds(pace_str):
    """配速（'5'20''）→ 秒/公里（320秒）"""
    minutes, seconds = map(int, pace_str.replace('"',"").split("'"))
    return minutes * 60 + seconds

def seconds_to_pace(seconds_float):
    """秒/公里（320.5秒）→ 配速（'5:20'）"""
    total_seconds = int(round(seconds_float))
    minutes = total_seconds // 60
    seconds = total_seconds % 60
    return f"{minutes}:{seconds:02d}"

def write_runner_record(records, ws, row_start, index, key, user_name):
    """
    records: 某一个人的多条记录
    ws: Excel worksheet
    row_start: 起始行号（建议奇数行），占2行（row_start 和 row_start+1）
    index: 序号
    """

    # print(row_start,user_name)

    days = ['周一', '周二', '周三', '周四', '周五', '周六', '周日']
    
    # 写入表头（仅首次）
    if row_start == 2:
        ws.cell(row=1, column=1, value='序号')
        ws.cell(row=1, column=2, value='组别')
        ws.cell(row=1, column=3, value='姓名')
        ws.cell(row=1, column=4, value='押金')
        ws.cell(row=1, column=5, value='本周剩余跑量')
        ws.cell(row=1, column=6, value='完成率')
        ws.cell(row=1, column=7, value='当前周跑量')
        for i, day in enumerate(days):
            ws.cell(row=1, column=8 + i, value=day)

    for col in range(1, 15):
        ws.cell(row=1, column=col).alignment = Alignment(vertical='center', horizontal='center')

    # 合并6列基本信息
    for col in range(1, 8):
        ws.merge_cells(start_row=row_start, start_column=col, end_row=row_start + 1, end_column=col)
        ws.cell(row=row_start, column=col).alignment = Alignment(vertical='center', horizontal='center')

    cell=ws.cell(row=row_start, column=1, value=index)
    fill_huang = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    cell.fill = fill_huang
    cell=ws.cell(row=row_start, column=2, value=key)
    color_str = "92D050"
    if key == '30公里':
        #浅蓝色
        color_str = "ADD8E6"
    elif key == '40公里':
        #浅橙色
        color_str = "FFA500"
    elif key == '50公里':
        #浅红色
        color_str = "FF6347"
    elif key == '70公里':
        #浅紫色
        color_str = "FF66FF"
    fill_lv = PatternFill(start_color=color_str, end_color=color_str, fill_type='solid')
    cell.fill = fill_lv

    cell.font = Font(bold=True)
    cell=ws.cell(row=row_start, column=3, value=user_name)
    cell.font = Font(bold=True)
    ws.cell(row=row_start, column=4, value=50)  # 押金
    cell=ws.cell(row=row_start, column=7, value=0)
    cell.font = Font(bold=True)
    ws.cell(row=row_start, column=7).alignment = Alignment(vertical='center', horizontal='center')

    # 颜色填充
    for col in range(8, 15):
        fill_lan = PatternFill(start_color='00FFFF', end_color='00FFFF', fill_type='solid')
        cell = ws.cell(row=row_start, column=col)
        cell.fill = fill_lan

        fill_zi = PatternFill(start_color='CC99FF', end_color='CC99FF', fill_type='solid')
        cell = ws.cell(row=row_start+1, column=col)
        cell.fill = fill_zi

    # 定义统一的边框样式（黑色细线）
    uniform_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000")
    )
    
    # 批量设置
    for col in range(1, 15):
        ws.cell(row=row_start, column=col).border = uniform_border
        ws.cell(row=row_start+1, column=col).border = uniform_border

    if len(records) == 0:
        return
    user = records[0]['userName']
    ws.cell(row=row_start, column=3, value=user)

    ws.cell(row=row_start, column=3).alignment = Alignment(vertical='center', horizontal='center')

    total_km = 0.0

    createTime_list = []
    for record in records:
        date = datetime.strptime(record['workoutDate'], "%Y-%m-%d")
        weekday = date.weekday()  # 0-6 → 周一到周日
        col = 8 + weekday
        current_km = round(record['distance'] / 1000, 2)
        current_pace = format_pace(record['avgPace'])  # 假设返回 '分:秒' 格式
        
        
        # 获取当前单元格的值
        km_cell = ws.cell(row=row_start, column=col)
        pace_cell = ws.cell(row=row_start + 1, column=col)
        
        if km_cell.value is None:
            # 首次写入：直接赋值
            new_km = current_km
            new_pace = current_pace
        elif record["createTime"] not  in createTime_list:
            # 累加公里数
            new_km = km_cell.value + current_km
            
            # 计算加权平均配速（单位：秒/公里）
            old_pace_seconds = pace_to_seconds(pace_cell.value)
            current_pace_seconds = pace_to_seconds(current_pace)
            
            total_time_seconds = (km_cell.value * old_pace_seconds) + (current_km * current_pace_seconds)
            avg_pace_seconds = total_time_seconds / new_km  # 加权平均
            
            new_pace = format_pace(seconds_to_pace(avg_pace_seconds))
        
        # 更新单元格
        km_cell.value = new_km
        km_cell.alignment = Alignment(vertical='center', horizontal='center')
        
        pace_cell.value = new_pace
        pace_cell.alignment = Alignment(vertical='center', horizontal='center')
        
        total_km += current_km  # 更新总公里数
        createTime_list.append(record["createTime"])


    # 本周剩余跑量
    delta = round(total_km-int(key.replace('公里','')), 2)
    value = f"超额 {delta}" if delta > 0 else f"{abs(delta)}"
    cell=ws.cell(row=row_start, column=5, value=value)
    cell.font = Font(bold=True)
    if delta >= 0:
        fill_huang = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        cell.fill = fill_huang
    ws.cell(row=row_start, column=5).alignment = Alignment(vertical='center', horizontal='center')

    # 完成率
    cell = ws.cell(row=row_start, column=6, value=f"{round(total_km/int(key.replace('公里',''))*100, 2)}%")
    ws.cell(row=row_start, column=6).alignment = Alignment(vertical='center', horizontal='center')
    cell.font = Font(bold=True)

    # 当前周跑量
    cell = ws.cell(row=row_start, column=7, value=round(total_km, 2))
    ws.cell(row=row_start, column=7).alignment = Alignment(vertical='center', horizontal='center')
    cell.font = Font(bold=True)

