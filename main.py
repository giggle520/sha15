from jielong import JIELONG_API
from ppt import PPT_API
import json
from prettytable import PrettyTable
from datetime import datetime
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from write_to_excel import write_runner_record
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side
from openpyxl.styles import Font
from datetime import datetime, timedelta
from config import EXCEL_PATH,TARGET_DATE

def get_week_range(target_date):
    """
    根据输入日期返回所在周的周一和周日
    
    Args:
        target_date (str/datetime): 输入日期（格式："YYYY-MM-DD" 或 datetime对象）
    Returns:
        tuple: (start_date, end_date) -> 均为字符串格式 "YYYY-MM-DD"
    """
    # 1. 统一转为datetime对象
    if isinstance(target_date, str):
        target_date = datetime.strptime(target_date, "%Y-%m-%d")
    
    # 2. 计算周一和周日
    monday = target_date - timedelta(days=target_date.weekday())
    sunday = monday + timedelta(days=6)
    
    # 3. 格式化为字符串
    return monday.strftime("%Y-%m-%d"), sunday.strftime("%Y-%m-%d")

def get_monday_of_week():
    """获取当前周的周一日期（格式：YYYYMMDD）"""
    today = datetime.now()
    # 计算本周周一（weekday()返回0-6，0是周一）
    monday = today - timedelta(days=today.weekday())
    return monday.strftime("%Y%m%d") + "周"

def format_pace(pace_str):
    if not pace_str or ':' not in pace_str:
        return pace_str
    minutes, seconds = pace_str.split(':')
    return f"{minutes}'" + f'{seconds}"'

def filter_by_date_range(records, start_date, end_date):
    """按日期区间筛选记录"""
    start = datetime.strptime(start_date, "%Y-%m-%d")
    end = datetime.strptime(end_date, "%Y-%m-%d")
    
    filtered = []
    for record in records:
        record_date = datetime.strptime(record['punchDate'], "%Y-%m-%d")
        if start <= record_date <= end:
            filtered.append(record)
    return filtered

def parse_running_records(json_data, start_date=None, end_date=None, ws=None, row_start=2, index=1, key=None, user_name=None):
    global wb

    """解析跑步记录（支持日期筛选）"""
    records = json_data.get('data', {}).get('records', [])

    # 日期筛选
    if start_date and end_date:
        records = filter_by_date_range(records, start_date, end_date)

    write_runner_record(records, ws, row_start=row, index=index, key=key, user_name=user_name)

    # wb.save(excel_path)

    # ws = wb.active

    # 设置列宽（单位：字符宽度）
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['E'].width = 15 
    ws.column_dimensions['G'].width = 15  

    for col in range(1, 15):
        cell=ws.cell(row=1, column=col)
        fill_lv = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
        cell.fill = fill_lv
        cell.font = Font(bold=True)

    
    # 定义统一的边框样式（黑色细线）
    uniform_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000")
    )
    
    # 批量设置
    for col in range(1, 15):
        ws.cell(row=1, column=col).border = uniform_border

    wb.save(excel_path)

    # 返回结构化数据
    return [
        {
            "name": record.get('userName'),
            "date": record.get('punchDate'),
            "distance_km": float(record.get('distance', 0)) / 1000,
            "pace": format_pace(record.get('avgPace')),
            "duration": record.get('duration')
        }
        for record in records
    ]

def map_username(user_name):
    # 定义特殊映射规则
    special_mapping = {
        '唐文超': '唐文超-24MEM',
        '张帅': '张帅-22MBA',
        '何文华': '何文华-2022MBA',
        '蔡宝岩': '蔡宝岩-23MEM',
        '叶橙': '叶橙-2023MBA',
        'syh': '石翊函'
    }
    
    # 检查用户名是否在特殊映射中
    if user_name in special_mapping:
        return special_mapping[user_name]
    else:
        # 如果不在特殊映射中，返回原用户名
        return user_name

def add_weekly_sheet(file_path, week_title=None):
    """自动按周创建新Sheet页"""
    # 1. 计算当前周标识（如 0615周）
    if week_title is None:
        week_title = get_monday_of_week()
    
    # 2. 加载Excel（若文件不存在则新建）
    try:
        wb = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # 删除默认Sheet
    
    # 3. 检查是否已存在该周Sheet
    if week_title in wb.sheetnames:
        wb.remove(wb[week_title])  # 删除旧Sheet
        print(f"删除现有Sheet: {week_title}")
    ws = wb.create_sheet(week_title)
    print(f"新建Sheet: {week_title}")

    return wb,ws

if __name__ == "__main__":
    #================================================================#

    excel_path = EXCEL_PATH

    target_date = TARGET_DATE
    start_date,end_date = get_week_range(target_date)
    week_title = start_date.replace("-", "") + "周"
    print(f"获取 {target_date} 所在周的日期范围: {start_date} 到 {end_date}")

    #================================================================#

    ppt_api = PPT_API()
    # 获取跑团成员数据
    all_users = ppt_api.get_users()
    print(f"\n共获取到 {len(all_users)} 条用户数据:")

    print("\n报名情况统计:")
    jielong_api = JIELONG_API()
    group_dict = jielong_api.get_jielong_info()

    wb,ws = add_weekly_sheet(excel_path, week_title)
    row = 2
    index = 1

    in_ppt_users = []
    not_in_ppt_users = []
    

    #区分在ppt的用户和不在ppt的用户
    for key in group_dict.keys():
        user_names = group_dict.get(key)
        for user_name in user_names:
            existing_users = {user["userName"] for user in all_users}
            if map_username(user_name) not in existing_users:
                not_in_ppt_users.append({"userName":user_name})
            else:
                in_ppt_users.append({"userName":user_name})

    for key in group_dict.keys():
        print("-" * 10)
        print(key)
        user_names = group_dict.get(key)
        for user_name in user_names:
            '''
            特殊mapping:
            唐文超 -> 唐文超-24MEM
            张帅 -> 张帅-22MBA
            何文华 -> 何文华-2022MBA
            蔡宝岩 -> 蔡宝岩-23MEM
            叶橙 -> 叶橙-2023MBA
            syh -> 石翊函
            '''
            for i, user in enumerate(all_users, 1): 
                if user['userName'] == map_username(user_name):
                    data = ppt_api.get_weekly_data(user['userId'], target_date=target_date)
                    filtered_data = parse_running_records(
                        data,
                        start_date=start_date,
                        end_date=end_date,
                        ws=ws,
                        row_start=row,
                        index=index,
                        key=key,
                        user_name=map_username(user_name)
                    )
                    row += 2
                    index += 1

            for item in not_in_ppt_users: 
                if item['userName'] == user_name:
                    parse_running_records(
                        {"data":{"records":[]}},
                        start_date=start_date,
                        end_date=end_date,
                        ws=ws,
                        row_start=row,
                        index=index,
                        key=key,
                        user_name=map_username(user_name)
                    )
                    row += 2
                    index += 1
                    break